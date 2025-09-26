# -*- coding: utf-8 -*-
r"""
MIDEA - Overview Operação Midea (Diário) — Streamlit

Ajustes:
- Carregar sempre a versão mais recente ao abrir (sem clicar em Atualizar)
- Cache-buster (nocache=) para furar cache do SharePoint/CDN/Streamlit
- Botão Atualizar continua funcionando
- width="stretch" no st.plotly_chart (com fallback automático)
- Para tabelas: use_container_width=True (st.dataframe)
- Forçar FREETIME (ou variantes) para texto, evitando ArrowTypeError
- Converter colunas com qualquer texto/objeto para pandas StringDtype (ex.: QTD com '???')
- Horário em America/Sao_Paulo (Brasília)
- DEVOLUÇÃO/CANCELAMENTO: validadores de linha para ignorar segundo cabeçalho/linhas “0”
- Legenda do gráfico em branco
"""

from __future__ import annotations
import os, re, io, time, tempfile, base64, datetime
import requests
import pandas as pd
import streamlit as st

# Tenta Plotly; se não, cai para Altair
HAS_PLOTLY = True
try:
    import plotly.graph_objects as go  # type: ignore
except Exception:
    HAS_PLOTLY = False

# ================== CONFIG ==================
# >>> LINK PERMANENTE DO ONEDRIVE (com ?download=1) <<<
DEFAULT_ONEDRIVE_URL = (
    "https://tecadi-my.sharepoint.com/:x:/g/personal/rafael_alves_tecadi_com_br/"
    "EaJshSFavb5Pv8z_dpW3ZWwBVhjuG3tFcYeSRUMWSEbYyg?download=1"
)
DEFAULT_ONEDRIVE_URL = os.environ.get("MIDEA_ONEDRIVE_URL", DEFAULT_ONEDRIVE_URL)

TARGET_SHEET_1 = "PROGRAMAÇÃO DIÁRIA"
TARGET_SHEET_2 = "PROCESSOS S.LEITURA"
HEADER_ROW = 5

AREAS = {
    TARGET_SHEET_1: {
        "RECEBIMENTO":  {"cols": ("A", "Q"),  "status_col": "J"},
        "EXPEDIACAO":   {"cols": ("U", "AD"), "status_col": "AB"},
        # DEVOLUÇÃO: AH:AU, status AR, denominador/validador em AK (DATA)
        "DEVOLUCAO":    {"cols": ("AH", "AU"), "status_col": "AR"},
        # CANCELAMENTO: AZ:BD, status BB, denominador/validador em BA (OS)
        "CANCELAMENTO": {"cols": ("AZ", "BD"), "status_col": "BB"},
    },
    TARGET_SHEET_2: {
        # FASTFOB -> STATUS é K (não J)
        "FASTFOB":     {"cols": ("A", "N"),  "status_col": "K"},   # filtra por ABS 'C' (Armador)
        "TRANSBORDO":  {"cols": ("P", "AA"), "status_col": "X"},   # filtra por ABS 'R' (Container)
    },
}

UI_NAME_1 = "Recebimento & Expedição"
UI_NAME_2 = "Transbordos & Fastfob"
UI_NAME_3 = "Devolução & Cancelamento"

BG_PATH_CANDIDATES = [
    r"C:\Users\felipe.nonato\Music\Projetos\10 - OV Midea\fundoapp.jpg",
    r"C:\Users\felipe.nonato\Music\Projetos\10 - OV Midea\fundoapp.jpeg",
    r"C:\Users\felipe.nonato\Music\Projetos\10 - OV Midea\fundoapp.png",
    os.path.join(os.path.dirname(__file__), "fundoapp.jpg"),
    os.path.join(os.path.dirname(__file__), "fundoapp.jpeg"),
    os.path.join(os.path.dirname(__file__), "fundoapp.png"),
]
LOGO_PATH_CANDIDATES = [
    r"C:\Users\felipe.nonato\Music\Projetos\10 - OV Midea\logomidea.png",
    os.path.join(os.path.dirname(__file__), "logomidea.png"),
]
FAVICON_CANDIDATES = [
    r"C:\Users\felipe.nonato\Music\Projetos\10 - OV Midea\favicon.ico",
    os.path.join(os.path.dirname(__file__), "favicon.ico"),
    os.path.join(os.path.dirname(__file__), "favicon.png"),
]

# ================== UTILS ==================
def is_finalizado(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    return "FINALIZAD" in str(val).strip().upper()

def is_filled(val) -> bool:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    s = str(val).strip()
    return s != "" and s.lower() != "none"

def clean_onedrive_url(url: str) -> str:
    if not url:
        return url
    if "download=1" not in url:
        if "?" in url:
            if re.search(r"[?&]download=", url):
                return url
            return url + "&download=1"
        return url + "?download=1"
    return url

def _guess_filename(url: str, resp: requests.Response | None) -> str:
    if resp:
        cd = resp.headers.get("Content-Disposition") or resp.headers.get("content-disposition") or ""
        m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^\";]+)"?', cd)
        if m:
            return os.path.basename(m.group(1))
    tail = os.path.basename(url.split("?", 1)[0].split("#", 1)[0])
    return tail or "arquivo.xlsx"

def _safe_ext_from_name(name: str) -> str:
    ext = os.path.splitext(name)[1].lower()
    return ext if ext in (".xlsx", ".xlsm") else ".xlsx"

@st.cache_data(show_spinner="Baixando arquivo do OneDrive…")
def _fetch_excel_bytes(url: str, refresh_key: int) -> bytes:
    final_url = clean_onedrive_url(url)
    sep = "&" if "?" in final_url else "?"
    final_url = f"{final_url}{sep}nocache={refresh_key}_{int(time.time())}"
    with requests.get(final_url, stream=True, timeout=60, allow_redirects=True) as r:
        r.raise_for_status()
        return r.content

def _bytes_to_tempfile(xbytes: bytes, name_hint: str = "arquivo.xlsx") -> str:
    tmpdir = tempfile.mkdtemp(prefix="midea_")
    ext = _safe_ext_from_name(name_hint)
    path = os.path.join(tmpdir, os.path.splitext(name_hint)[0] + ext)
    with open(path, "wb") as f:
        f.write(xbytes)
    return path

def _last_used_row_iter(ws, start_col_idx, end_col_idx, header_row):
    last = header_row
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=header_row+1, min_col=start_col_idx, max_col=end_col_idx, values_only=True),
        start=header_row+1,
    ):
        if any(c not in (None, "") for c in row):
            last = r_idx
    return last

def read_range_df(path: str, sheet_name: str, col_from: str, col_to: str, header_row: int = HEADER_ROW) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        raise RuntimeError("Arquivo .xls detectado. Salve como .xlsx/.xlsm.")
    import openpyxl
    from openpyxl.utils import column_index_from_string
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_vba=(ext == ".xlsm"))
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba '{sheet_name}' não encontrada.")
    ws = wb[sheet_name]
    start = column_index_from_string(col_from)
    end   = column_index_from_string(col_to)
    headers = [ws.cell(row=header_row, column=c).value for c in range(start, end + 1)]
    headers = [str(h) if h not in (None, "") else f"COL_{i+1}" for i, h in enumerate(headers)]
    last_row = _last_used_row_iter(ws, start, end, header_row)
    if last_row <= header_row:
        wb.close()
        return pd.DataFrame(columns=headers)
    data = [list(row) for row in ws.iter_rows(min_row=header_row+1, max_row=last_row, min_col=start, max_col=end, values_only=True)]
    wb.close()
    return pd.DataFrame(data, columns=headers).dropna(how="all")

def _rel_idx_from_abs_letter(abs_letter: str, area_cols: tuple[str, str]) -> int | None:
    from openpyxl.utils import column_index_from_string
    start = column_index_from_string(area_cols[0])
    end   = column_index_from_string(area_cols[1])
    abs_i = column_index_from_string(abs_letter)
    if abs_i < start or abs_i > end:
        return None
    return abs_i - start

# ---- Validadores específicos para evitar "segundo cabeçalho" ----
def _is_date_like(val) -> bool:
    """True se for data válida (datetime/date ou string que vira data)."""
    if isinstance(val, (datetime.date, datetime.datetime, pd.Timestamp)):
        return True
    s = str(val).strip()
    if s in ("", "0", "TOTAL"):
        return False
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return pd.notna(dt)
    except Exception:
        return False

_os_re = re.compile(r"^[A-Za-z]\d{3,}$")
def _looks_like_os(val) -> bool:
    """True para OS no formato típico (ex.: A37750)."""
    if val is None:
        return False
    s = str(val).strip()
    if s in ("", "0", "TOTAL"):
        return False
    return bool(_os_re.match(s))

def count_area_status_filtered(
    path: str,
    sheet_name: str,
    area_key: str,
    required_abs_col: str | None = None,
    required_validator=None,
):
    """Lê o intervalo da área e aplica:
       - filtro por required_abs_col (se fornecido), usando required_validator (se fornecido);
       - depois contabiliza STATUS (finalizado vs em aberto).
    """
    from openpyxl.utils import column_index_from_string

    area = AREAS[sheet_name][area_key]
    df = read_range_df(path, sheet_name, area["cols"][0], area["cols"][1], header_row=HEADER_ROW)

    start_abs = column_index_from_string(area["cols"][0])
    status_abs = column_index_from_string(area["status_col"])
    rel_status = status_abs - start_abs
    if rel_status < 0 or rel_status >= len(df.columns):
        cand = [i for i, h in enumerate(df.columns) if h and str(h).strip().lower().startswith("status")]
        rel_status = cand[0] if cand else None

    if required_abs_col:
        rel_req = _rel_idx_from_abs_letter(required_abs_col, area["cols"])
        if rel_req is not None and rel_req < len(df.columns):
            ser = df.iloc[:, rel_req]
            if required_validator is not None:
                mask = ser.apply(required_validator)
            else:
                mask = ser.apply(is_filled)
            df = df[mask].copy()

    total = len(df.index)
    finalizados = 0
    if total > 0 and rel_status is not None and rel_status < len(df.columns):
        s = df.iloc[:, rel_status]
        finalizados = int(sum(is_finalizado(x) for x in s.values))
    em_aberto = total - finalizados
    return total, finalizados, em_aberto, df

# ================== THEME & CSS ==================
def _mime_from_path(p: str) -> str:
    ext = os.path.splitext(p)[1].lower()
    if ext in (".jpg", ".jpeg"): return "image/jpeg"
    if ext == ".png":            return "image/png"
    if ext == ".ico":            return "image/x-icon"
    return "image/jpeg"

def _inject_theme_and_background():
    bg_path = next((p for p in BG_PATH_CANDIDATES if os.path.exists(p)), None)
    logo_path = next((p for p in LOGO_PATH_CANDIDATES if os.path.exists(p)), None)

    bg_css = ""
    if bg_path:
        with open(bg_path, "rb") as f:
            bg_b64 = base64.b64encode(f.read()).decode()
        mime = _mime_from_path(bg_path)
        bg_css = f"""
        .stApp {{
            background: url('data:{mime};base64,{bg_b64}') no-repeat center center / cover fixed;
        }}
        .stApp::before {{
            content: ""; position: fixed; inset: 0;
            background: rgba(0,0,0,0.28);
            z-index: -1; pointer-events: none;
        }}
        """

    logo_css = ""
    if logo_path:
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        logo_css = f"""
        .stApp::after {{
            content: ""; position: fixed; top: 10px; right: 16px; z-index: 999;
            width: 180px; height: 64px;
            background: url('data:image/png;base64,{logo_b64}') no-repeat right top / contain;
            pointer-events: none; opacity: .95;
        }}
        """

    st.markdown(f"""
    <style>
    div[data-testid="stToolbar"], div[data-testid="stStatusWidget"], div[data-testid="stDecoration"] {{ display: none !important; }}
    #MainMenu {{ visibility: hidden; }}
    header {{ visibility: hidden; height: 0; }}
    footer {{ visibility: hidden; }}

    {bg_css}
    {logo_css}

    .block-container, .block-container *:not([data-testid="stDataFrame"] *),
    h1, h2, h3, h4, h5, h6, p, label, span {{ color: #fff !important; }}

    div[data-testid="stDataFrame"] * {{ color: #111 !important; }}

    .block-container {{ padding-top: 0.6rem !important; }}

    .card {{
        background: rgba(255,255,255,0.08);
        border: 3px solid rgba(0, 190, 255, 0.55);
        border-radius: 28px; padding: 18px 20px 14px 20px;
        box-shadow: 0 6px 20px rgba(0,0,0,0.25);
        backdrop-filter: blur(2px);
    }}
    .card .pill {{
        display:inline-block; padding: 8px 18px; border-radius: 999px;
        border: 1px solid rgba(0, 190, 255, 0.75);
        background: linear-gradient(180deg, rgba(0, 190, 255, 0.18), rgba(0, 190, 255, 0.08));
        color:#e9faff; font-weight:700; letter-spacing:.2px; margin-bottom: 6px;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.25);
    }}
    div[data-testid="stMetricValue"] {{
        font-size: 48px !important; line-height: 1.0 !important; margin-top: 2px !important;
    }}

    .stButton button {{
      background: rgba(0, 120, 255, 0.90) !important;
      color: #ffffff !important;
      border: 1px solid rgba(255,255,255,0.5) !important;
      border-radius: 14px !important;
      padding: 8px 14px !important;
      box-shadow: 0 3px 12px rgba(0,0,0,0.25) !important;
    }}
    .stButton button:hover {{
      filter: brightness(1.06);
      transform: translateY(-1px);
    }}

    .stSelectbox label {{ color:#fff !important; font-weight:700; }}
    .stSelectbox [data-baseweb="select"] > div {{
      background:#0F2846 !important; border:2px solid #00BFFF !important; border-radius:14px !important;
      min-height:52px; box-shadow:0 3px 12px rgba(0,0,0,.25);
    }}
    .stSelectbox [data-baseweb="select"] * {{ color:#fff !important; }}
    .stSelectbox [data-baseweb="select"] svg {{ fill:#fff !important; }}
    ul[role="listbox"] {{ background:#0F2846 !important; color:#fff !important; border:1px solid #00BFFF !important; }}
    ul[role="listbox"] li {{ color:#fff !important; }}
    ul[role="listbox"] li:hover {{ background:rgba(0,191,255,.15) !important; }}
    .stSelectbox [data-baseweb="select"] > div > div {{ font-size:1.05rem; font-weight:700; }}

    div[data-testid="stDownloadButton"] button,
    div[data-testid="stDownloadButton"] > div > button {{
        background: rgba(0, 120, 255, 0.9) !important; color: #fff !important;
        border: 1px solid rgba(255,255,255,0.5) !important; border-radius: 14px !important;
        padding: 8px 14px !important; box-shadow: 0 3px 12px rgba(0,0,0,0.25) !important;
    }}
    div[data-testid="stDownloadButton"] button:hover {{ filter: brightness(1.06); transform: translateY(-1px); }}
    </style>
    """, unsafe_allow_html=True)

def card_metric(title: str, value: int | float | str, legend: str = ""):
    with st.container():
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f"<div class='pill'>{title}</div>", unsafe_allow_html=True)
        st.metric(label=" ", value=value)
        if legend:
            st.markdown(f"<div style='margin-top:-2px;font-size:0.95rem;opacity:.92'>{legend}</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

def build_overview_chart(labels, finalizados, em_aberto, totais):
    cor_andamento = "#425F96"
    cor_finalizado = "#177833"
    if HAS_PLOTLY:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=labels, y=em_aberto,   name="Em andamento",
                             marker=dict(color=cor_andamento),
                             hovertemplate="Em andamento: %{y}<extra></extra>"))
        fig.add_trace(go.Bar(x=labels, y=finalizados, name="Finalizado",
                             marker=dict(color=cor_finalizado),
                             hovertemplate="Finalizado: %{y}<extra></extra>"))
        fig.add_trace(go.Scatter(x=labels, y=totais, mode="text",
                                 text=[str(t) for t in totais],
                                 textposition="top center",
                                 textfont=dict(size=18, color="#FFFFFF"),
                                 hoverinfo="skip", showlegend=False))
        fig.update_layout(
            barmode="stack",
            height=420,
            margin=dict(l=10, r=10, t=30, b=10),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#FFFFFF"),
            legend=dict(
                orientation="h", yanchor="bottom", y=1.04, xanchor="right", x=1,
                font=dict(color="#FFFFFF")
            ),
        )
        fig.update_xaxes(tickfont=dict(color="#FFFFFF"), color="#FFFFFF",
                         gridcolor="rgba(255,255,255,0.15)")
        fig.update_yaxes(tickfont=dict(color="#FFFFFF"), color="#FFFFFF",
                         gridcolor="rgba(255,255,255,0.15)")
        return fig, "plotly"
    else:
        import altair as alt
        df = pd.DataFrame({
            "Operacao": labels, "Finalizado": finalizados, "Em andamento": em_aberto, "Total": totais,
        })
        df_long = df.melt(id_vars=["Operacao", "Total"],
                          value_vars=["Em andamento", "Finalizado"],
                          var_name="Status", value_name="Valor")
        chart = (
            alt.Chart(df_long).mark_bar()
            .encode(
                x=alt.X("Operacao:N", axis=alt.Axis(labelColor="white", title=None)),
                y=alt.Y("Valor:Q", stack="zero", axis=alt.Axis(labelColor="white", title=None)),
                color=alt.Color("Status:N",
                                scale=alt.Scale(domain=["Em andamento", "Finalizado"],
                                                range=["#425F96", "#177833"]),
                                legend=alt.Legend(labelColor="white", titleColor="white"))
            )
            .properties(height=380).configure_view(stroke=None)
        )
        txt = alt.Chart(df).mark_text(color="white", dy=-6, fontSize=18)\
            .encode(x=alt.X("Operacao:N"), y=alt.Y("Total:Q"), text="Total:Q")
        return (chart + txt).configure_axis(grid=False, labelColor="white", titleColor="white"), "altair"

# ================== APP ==================
page_icon_path = next((p for p in FAVICON_CANDIDATES if os.path.exists(p)), None)
st.set_page_config(page_title="MIDEA - Overview (Tecadi)", page_icon=page_icon_path, layout="wide")
_inject_theme_and_background()

if "refresh_counter" not in st.session_state:
    st.session_state.refresh_counter = int(time.time())
if "last_updated" not in st.session_state:
    st.session_state.last_updated = None
if "fixed_url" not in st.session_state:
    st.session_state.fixed_url = DEFAULT_ONEDRIVE_URL.strip()

st.title("MIDEA — Overview Operação (Diário)")

try:
    from zoneinfo import ZoneInfo
    tz_br = ZoneInfo("America/Sao_Paulo")
except Exception:
    tz_br = None

top_l, top_r = st.columns([0.80, 0.20])
with top_r:
    if st.button("🔄 Atualizar dados", type="primary", help="Recarrega o Excel do OneDrive"):
        st.session_state.refresh_counter += 1
        st.rerun()
    if st.session_state.last_updated:
        if tz_br:
            st.caption(f"Atualizado: {st.session_state.last_updated.astimezone(tz_br):%d/%m %H:%M:%S} (Brasília)")
        else:
            st.caption(f"Atualizado: {st.session_state.last_updated:%d/%m %H:%M:%S}")

st.caption("Versão v1.1 — Overview & Exportações")

if not st.session_state.fixed_url:
    st.subheader("Informe o link do OneDrive para iniciar")
    link = st.text_input("Link do OneDrive (compartilhado)", placeholder="cole aqui o link do arquivo…")
    if st.button("Carregar do OneDrive", type="primary"):
        if not link.strip():
            st.error("Informe o link do OneDrive.")
        else:
            st.session_state.fixed_url = link.strip()
            st.session_state.refresh_counter += 1
            st.rerun()
    st.stop()

try:
    xbytes = _fetch_excel_bytes(st.session_state.fixed_url, st.session_state.refresh_counter)
    name_hint = _guess_filename(st.session_state.fixed_url, None)
    local_path = _bytes_to_tempfile(xbytes, name_hint=name_hint)
    st.session_state.last_updated = datetime.datetime.now(datetime.timezone.utc)
except Exception as e:
    st.error("Erro ao baixar a planilha do OneDrive. Verifique o link/compartilhamento.")
    st.exception(e)
    st.stop()

# Leitura dos dados
try:
    r_tot, r_fin, r_ab, df_rec   = count_area_status_filtered(local_path, TARGET_SHEET_1, "RECEBIMENTO")
    e_tot, e_fin, e_ab, df_exp   = count_area_status_filtered(local_path, TARGET_SHEET_1, "EXPEDIACAO")
    f_tot, f_fin, f_ab, df_fast  = count_area_status_filtered(local_path, TARGET_SHEET_2, "FASTFOB",    required_abs_col="C")
    t_tot, t_fin, t_ab, df_tran  = count_area_status_filtered(local_path, TARGET_SHEET_2, "TRANSBORDO", required_abs_col="R")

    # >>> DEVOLUÇÃO/CANCELAMENTO com validadores <<<
    d_tot, d_fin, d_ab, df_devol = count_area_status_filtered(
        local_path, TARGET_SHEET_1, "DEVOLUCAO", required_abs_col="AK", required_validator=_is_date_like
    )
    c_tot, c_fin, c_ab, df_canc  = count_area_status_filtered(
        local_path, TARGET_SHEET_1, "CANCELAMENTO", required_abs_col="BA", required_validator=_looks_like_os
    )
except Exception as e:
    st.error("Erro ao ler planilha. Verifique as abas/intervalos.")
    st.exception(e)
    st.stop()

# Gráfico resumo (4 barras originais)
st.subheader("Resumo — Andamento das Operações")
labels = ["Recebimento", "Expedição", "Fastfob", "Transbordo s/ Leitura"]
finalizados = [r_fin, e_fin, f_fin, t_fin]
em_aberto   = [r_ab,  e_ab,  f_ab,  t_ab]
totais      = [r_tot, e_tot, f_tot, t_tot]

chart, kind = build_overview_chart(labels, finalizados, em_aberto, totais)
if kind == "plotly":
    # usa width="stretch"; se a versão do Streamlit não suportar, cai para use_container_width=True
    try:
        st.plotly_chart(chart, width="stretch", config={"displayModeBar": False})
    except TypeError:
        st.plotly_chart(chart, use_container_width=True, config={"displayModeBar": False})
else:
    st.altair_chart(chart, use_container_width=True)

st.divider()

# ---- Arrow-friendly DataFrames ----
def _coerce_freetime_to_str(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if str(col).strip().upper() in {"FREETIME", "FREE TIME", "FREE_TIME", "FREETIME (H)", "FREETIME(H)"}:
            df[col] = df[col].astype(str)
    return df

def _df_for_display(df: pd.DataFrame) -> pd.DataFrame:
    from pandas.api.types import is_object_dtype
    df = _coerce_freetime_to_str(df.copy())
    for c in df.columns:
        s = df[c]
        if is_object_dtype(s) or s.map(lambda x: isinstance(x, str)).any():
            try:
                df[c] = s.astype("string[python]")
            except Exception:
                df[c] = s.astype(str)
    return df

df_rec_disp   = _df_for_display(df_rec)
df_exp_disp   = _df_for_display(df_exp)
df_fast_disp  = _df_for_display(df_fast)
df_tran_disp  = _df_for_display(df_tran)
df_devol_disp = _df_for_display(df_devol)
df_canc_disp  = _df_for_display(df_canc)

# Detalhes por grupo
opt = st.selectbox("Selecione o grupo", [UI_NAME_1, UI_NAME_2, UI_NAME_3], index=0)
st.divider()

if opt == UI_NAME_1:
    c1, c2 = st.columns(2)
    with c1:
        card_metric("RECEBIMENTOS (total)", r_tot, legend=f"Finalizados: {r_fin} · Em andamento: {r_ab}")
    with c2:
        card_metric("EXPEDIÇÕES (total)", e_tot, legend=f"Finalizados: {e_fin} · Em andamento: {e_ab}")

    st.markdown("**Recebimento (A:Q)**")
    st.dataframe(df_rec_disp, use_container_width=True)

    st.markdown("**Expedição (U:AD)**")
    st.dataframe(df_exp_disp, use_container_width=True)

    col_a, col_b, _ = st.columns([1,1,2])
    with col_a:
        if not df_rec_disp.empty:
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as xw:
                df_rec_disp.to_excel(xw, sheet_name="RECEBIMENTO", index=False)
            st.download_button("Baixar RECEBIMENTO (.xlsx)", data=bio.getvalue(),
                               file_name="recebimento.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col_b:
        if not df_exp_disp.empty:
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as xw:
                df_exp_disp.to_excel(xw, sheet_name="EXPEDICAO", index=False)
            st.download_button("Baixar EXPEDIÇÃO (.xlsx)", data=bio.getvalue(),
                               file_name="expedicao.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

elif opt == UI_NAME_2:
    c1, c2 = st.columns(2)
    with c1:
        card_metric("FASTFOB (total)", f_tot, legend=f"Finalizados: {f_fin} · Em andamento: {f_ab}")
    with c2:
        card_metric("TRANSBORDO S/ LEITURA (total)", t_tot, legend=f"Finalizados: {t_fin} · Em andamento: {t_ab}")

    st.markdown("**FASTFOB (A:N)**")
    st.dataframe(df_fast_disp, use_container_width=True)

    st.markdown("**TRANSBORDO (P:AA)**")
    st.dataframe(df_tran_disp, use_container_width=True)

    col_c, col_d, _ = st.columns([1,1,2])
    with col_c:
        if not df_fast_disp.empty:
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as xw:
                df_fast_disp.to_excel(xw, sheet_name="FASTFOB", index=False)
            st.download_button("Baixar FASTFOB (.xlsx)", data=bio.getvalue(),
                               file_name="fastfob.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col_d:
        if not df_tran_disp.empty:
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as xw:
                df_tran_disp.to_excel(xw, sheet_name="TRANSBORDO", index=False)
            st.download_button("Baixar TRANSBORDO (.xlsx)", data=bio.getvalue(),
                               file_name="transbordo.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:  # UI_NAME_3 -> Devolução & Cancelamento
    c1, c2 = st.columns(2)
    with c1:
        card_metric("DEVOLUÇÕES (total)", d_tot, legend=f"Finalizados: {d_fin} · Em andamento: {d_ab}")
    with c2:
        card_metric("CANCELAMENTOS (total)", c_tot, legend=f"Finalizados: {c_fin} · Em andamento: {c_ab}")

    st.markdown("**DEVOLUÇÃO (AH:AU)**")
    st.dataframe(df_devol_disp, use_container_width=True)

    st.markdown("**CANCELAMENTO (AZ:BD)**")
    st.dataframe(df_canc_disp, use_container_width=True)

    col_e, col_f, _ = st.columns([1,1,2])
    with col_e:
        if not df_devol_disp.empty:
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as xw:
                df_devol_disp.to_excel(xw, sheet_name="DEVOLUCAO", index=False)
            st.download_button("Baixar DEVOLUÇÃO (.xlsx)", data=bio.getvalue(),
                               file_name="devolucao.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col_f:
        if not df_canc_disp.empty:
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as xw:
                df_canc_disp.to_excel(xw, sheet_name="CANCELAMENTO", index=False)
            st.download_button("Baixar CANCELAMENTO (.xlsx)", data=bio.getvalue(),
                               file_name="cancelamento.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Rodapé fixo
st.caption("© Tecadi — versão v1.1. Autor: FG • Contato: felipe.nonato@tecadi.com.br")
